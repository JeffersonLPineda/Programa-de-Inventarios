from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem,
    QPushButton, QMessageBox, QHBoxLayout, QDateEdit, QFileDialog, QComboBox
)
from PyQt6.QtGui import QIcon
from PyQt6.QtCore import QDate, Qt
import sqlite3
import database as db
import os
import sys
import re

# Exportaciones
from openpyxl import Workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet


def safe_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def resource_path(relative_path):
    base_path = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base_path, relative_path)


class KardexWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Kardex de Ventas")
        self.setMinimumSize(1200, 700)
        try:
            self.setWindowIcon(QIcon(resource_path("mainlogo.ico")))
        except Exception:
            pass

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout()

        # Selección de rango de fechas + método
        top_layout = QHBoxLayout()
        top_layout.addWidget(QLabel("Fecha inicio:"))
        self.fecha_inicio = QDateEdit()
        self.fecha_inicio.setCalendarPopup(True)
        self.fecha_inicio.setDate(QDate.currentDate().addDays(-30))
        top_layout.addWidget(self.fecha_inicio)

        top_layout.addWidget(QLabel("Fecha fin:"))
        self.fecha_fin = QDateEdit()
        self.fecha_fin.setCalendarPopup(True)
        self.fecha_fin.setDate(QDate.currentDate())
        top_layout.addWidget(self.fecha_fin)

        top_layout.addWidget(QLabel("Método:"))
        self.metodo_combo = QComboBox()
        self.metodo_combo.addItems(["PMP (Promedio Ponderado)", "PEPS (FIFO)", "UEPS (LIFO)"])
        top_layout.addWidget(self.metodo_combo)

        layout.addLayout(top_layout)

        # Botón mostrar Kardex
        btn_mostrar = QPushButton("Mostrar Kardex")
        btn_mostrar.clicked.connect(self.mostrar_kardex)
        layout.addWidget(btn_mostrar)

        # Botones exportación
        export_layout = QHBoxLayout()
        btn_excel = QPushButton("Exportar a Excel")
        btn_excel.clicked.connect(self.exportar_excel)
        export_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("Exportar a PDF")
        btn_pdf.clicked.connect(self.exportar_pdf)
        export_layout.addWidget(btn_pdf)

        layout.addLayout(export_layout)

        # Tabla Kardex
        self.kardex_table = QTableWidget()
        layout.addWidget(self.kardex_table)

        central.setLayout(layout)

    def mostrar_kardex(self):
        fecha_inicio = self.fecha_inicio.date().toString("yyyy-MM-dd")
        fecha_fin = self.fecha_fin.date().toString("yyyy-MM-dd")
        metodo_label = self.metodo_combo.currentText()
        if metodo_label.startswith("PMP"):
            metodo = "PMP"
        elif metodo_label.startswith("PEPS"):
            metodo = "PEPS"
        else:
            metodo = "UEPS"

        with sqlite3.connect(db.DB_NAME) as conn:
            cursor = conn.cursor()

            # --- Obtener inventarios (lotes reales) y ventas hasta fecha_fin ---
            cursor.execute("""
                SELECT i.fecha_compra AS fecha, i.id AS id_inventario, i.id_compra, i.id_producto, p.nombre, i.cantidad, i.precio_unitario
                FROM inventarios i
                JOIN productos p ON p.id_producto = i.id_producto
                JOIN compras c ON c.id_compra = i.id_compra
                WHERE i.fecha_compra <= ?
                ORDER BY i.fecha_compra ASC, i.id ASC
            """, (fecha_fin,))
            inventarios_todo = cursor.fetchall()

            cursor.execute("""
                SELECT v.fecha AS fecha, v.id_venta, dv.id_producto, p.nombre, dv.cantidad, dv.precio_unitario
                FROM ventas v
                JOIN detalle_ventas dv ON dv.id_venta = v.id_venta
                JOIN productos p ON p.id_producto = dv.id_producto
                WHERE v.fecha <= ?
                ORDER BY v.fecha ASC, v.id_venta ASC
            """, (fecha_fin,))
            ventas_todo = cursor.fetchall()

            if not inventarios_todo and not ventas_todo:
                QMessageBox.information(self, "Info", "No hay movimientos hasta la fecha seleccionada.")
                self.kardex_table.clear()
                self.kardex_table.setRowCount(0)
                self.kardex_table.setColumnCount(0)
                return

            # --- Construir eventos (compras como inventarios, y ventas) ---
            eventos = []
            for fecha, id_inventario, id_compra, id_producto, nombre, cantidad, precio in inventarios_todo:
                eventos.append({
                    "fecha": fecha,
                    "tipo": "compra",
                    "ref": id_inventario,
                    "id_inventario": id_inventario,
                    "id_compra": id_compra,
                    "producto_id": id_producto,
                    "producto": nombre,
                    "cantidad": safe_int(cantidad),
                    "precio": safe_float(precio)
                })
            for fecha, id_venta, id_producto, nombre, cantidad, precio in ventas_todo:
                eventos.append({
                    "fecha": fecha,
                    "tipo": "venta",
                    "ref": id_venta,
                    "producto_id": id_producto,
                    "producto": nombre,
                    "cantidad": safe_int(cantidad),
                    "precio": safe_float(precio)
                })

            eventos.sort(key=lambda e: (e["fecha"], 0 if e["tipo"] == "compra" else 1))

            # --- Estructuras por producto ---
            product_lots = {}   # pid -> [ { id_inventario, cantidad, precio, fecha }, ... ]
            product_prom = {}   # PMP: pid -> { cantidad, precio_prom }
            productos_seen = set()

            # --- Procesar eventos ANTES de fecha_inicio para inventario inicial ---
            for ev in eventos:
                if ev["fecha"] >= fecha_inicio:
                    break
                pid = ev["producto_id"]
                productos_seen.add(pid)
                if ev["tipo"] == "compra":
                    if metodo == "PMP":
                        prog = product_prom.get(pid, {"cantidad": 0, "precio_prom": 0.0})
                        q0, p0 = prog["cantidad"], prog["precio_prom"]
                        q1, p1 = ev["cantidad"], ev["precio"]
                        if q0 + q1 > 0:
                            nuevo_prom = ((q0 * p0) + (q1 * p1)) / (q0 + q1)
                        else:
                            nuevo_prom = 0.0
                        product_prom[pid] = {"cantidad": q0 + q1, "precio_prom": nuevo_prom}
                    else:
                        product_lots.setdefault(pid, []).append({
                            "id_inventario": ev["id_inventario"],
                            "cantidad": ev["cantidad"],
                            "precio": ev["precio"],
                            "fecha": ev["fecha"]
                        })
                else:
                    qv = ev["cantidad"]
                    if metodo == "PMP":
                        prog = product_prom.get(pid, {"cantidad": 0, "precio_prom": 0.0})
                        tomado = min(prog["cantidad"], qv)
                        prog["cantidad"] = max(prog["cantidad"] - tomado, 0)
                        product_prom[pid] = prog
                    else:
                        product_lots.setdefault(pid, [])
                        remaining = qv
                        if metodo == "UEPS":
                            idx = 0
                            while remaining > 0 and idx < len(product_lots[pid]):
                                lot = product_lots[pid][idx]
                                avail = safe_int(lot["cantidad"])
                                if avail <= 0:
                                    idx += 1
                                    continue
                                take = min(avail, remaining)
                                lot["cantidad"] = avail - take
                                remaining -= take
                                if lot["cantidad"] <= 0:
                                    idx += 1
                        else:  # PEPS consume entrada más reciente
                            idx = len(product_lots[pid]) - 1
                            while remaining > 0 and idx >= 0:
                                lot = product_lots[pid][idx]
                                avail = safe_int(lot["cantidad"])
                                if avail <= 0:
                                    idx -= 1
                                    continue
                                take = min(avail, remaining)
                                lot["cantidad"] = avail - take
                                remaining -= take
                                if lot["cantidad"] <= 0:
                                    idx -= 1

            # --- Preparar la tabla UI (añadida columna "ID" y ajustar ancho) ---
            columnas = [
                "Fecha", "Tipo", "Producto", "ID",
                "Cantidad", "Precio Unitario", "Valor Total",
                "Cantidad", "Precio Unitario", "Valor Total",
                "Cantidad", "Precio Unitario", "Valor Total"
            ]
            grupos = [
                "", "", "", "",
                "Entradas", "Entradas", "Entradas",
                "Salidas", "Salidas", "Salidas",
                "Inventario Final", "Inventario Final", "Inventario Final"
            ]

            self.kardex_table.clear()
            self.kardex_table.setColumnCount(len(columnas))
            self.kardex_table.setRowCount(2)
            for col, texto in enumerate(grupos):
                item = QTableWidgetItem(texto)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.kardex_table.setItem(0, col, item)
            for col, texto in enumerate(columnas):
                item = QTableWidgetItem(texto)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                # si es la columna de ID, poner el título en la fila 0 (porque está span-eada)
                if col == 3:
                    self.kardex_table.setItem(0, 3, QTableWidgetItem("ID Inventario"))
                    self.kardex_table.item(0, 3).setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                else:
                    self.kardex_table.setItem(1, col, item)


            # Cols spans
            self.kardex_table.setSpan(0, 0, 2, 1)
            self.kardex_table.setSpan(0, 1, 2, 1)
            self.kardex_table.setSpan(0, 2, 2, 1)
            self.kardex_table.setSpan(0, 3, 2, 1)

            def merge_group(start_col, span, text):
                self.kardex_table.setSpan(0, start_col, 1, span)
                itm = self.kardex_table.item(0, start_col)
                if itm:
                    itm.setText(text)
                else:
                    self.kardex_table.setItem(0, start_col, QTableWidgetItem(text))

            merge_group(4, 3, "Entradas")
            merge_group(7, 3, "Salidas")
            merge_group(10, 3, "Inventario Final")

            # ajustar ancho de la columna ID para que no sea muy grande
            try:
                self.kardex_table.setColumnWidth(3, 70)  # 70 px
            except Exception:
                pass

            row_idx = 2

            # Helper: para PEPS/UEPS devuelve lista de lots restantes; para PMP devuelve (cantidad, precio_prom)
            def remaining_lots(pid):
                if metodo == "PMP":
                    prog = product_prom.get(pid, {"cantidad": 0, "precio_prom": 0.0})
                    return prog["cantidad"], prog["precio_prom"]
                else:
                    lots = product_lots.get(pid, [])
                    return [(l["cantidad"], l["precio"], l["id_inventario"], l["fecha"]) for l in lots if safe_int(l["cantidad"]) != 0]

            # --- Recorrer eventos dentro del rango y generar filas ---
            for ev in eventos:
                if ev["fecha"] < fecha_inicio:
                    continue
                if ev["fecha"] > fecha_fin:
                    break

                pid = ev["producto_id"]
                nombre = ev["producto"]
                productos_seen.add(pid)

                if ev["tipo"] == "compra":
                    if metodo == "PMP":
                        prog = product_prom.get(pid, {"cantidad": 0, "precio_prom": 0.0})
                        q0, p0 = prog["cantidad"], prog["precio_prom"]
                        q1, p1 = ev["cantidad"], ev["precio"]
                        if q0 + q1 > 0:
                            nuevo_prom = ((q0 * p0) + (q1 * p1)) / (q0 + q1)
                        else:
                            nuevo_prom = 0.0
                        product_prom[pid] = {"cantidad": q0 + q1, "precio_prom": nuevo_prom}

                        total_after = product_prom[pid]["cantidad"]
                        avg_after = product_prom[pid]["precio_prom"]

                        compra_row = [
                            ev["fecha"], "Compra", str(nombre), str(ev["ref"]),
                            str(ev["cantidad"]), f"{ev['precio']:.2f}", f"{(ev['cantidad'] * ev['precio']):.2f}",
                            "-", "-", "-",
                            str(total_after), f"{avg_after:.2f}", f"{(total_after * avg_after):.2f}"
                        ]
                        self.kardex_table.insertRow(row_idx)
                        for col, val in enumerate(compra_row):
                            self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
                        row_idx += 1
                    else:
                        product_lots.setdefault(pid, []).append({
                            "id_inventario": ev["id_inventario"],
                            "cantidad": ev["cantidad"],
                            "precio": ev["precio"],
                            "fecha": ev["fecha"]
                        })
                        compra_row = [
                            ev["fecha"], "Compra", str(nombre), str(ev["id_inventario"]),
                            str(ev["cantidad"]), f"{ev['precio']:.2f}", f"{(ev['cantidad'] * ev['precio']):.2f}",
                            "-", "-", "-",
                            str(ev["cantidad"]), f"{ev['precio']:.2f}", f"{(ev['cantidad'] * ev['precio']):.2f}"
                        ]
                        self.kardex_table.insertRow(row_idx)
                        for col, val in enumerate(compra_row):
                            self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
                        row_idx += 1

                else:  # venta dentro del rango
                    q_venta = ev["cantidad"]

                    if metodo == "PMP":
                        prog = product_prom.get(pid, {"cantidad": 0, "precio_prom": 0.0})
                        disponible = prog["cantidad"]
                        if q_venta <= disponible:
                            prog["cantidad"] = disponible - q_venta
                            shortage = 0
                        else:
                            shortage = q_venta - disponible
                            prog["cantidad"] = 0
                        product_prom[pid] = prog

                        total_after = prog["cantidad"]
                        precio_prom = prog["precio_prom"]

                        venta_row = [
                            ev["fecha"], "Venta", str(nombre), "-",
                            "-", "-", "-",
                            str(q_venta), f"{ev['precio']:.2f}", f"{(q_venta * ev['precio']):.2f}",
                            str(total_after if shortage == 0 else -shortage), f"{precio_prom:.2f}", f"{( (total_after if total_after>0 else 0) * precio_prom):.2f}"
                        ]
                        self.kardex_table.insertRow(row_idx)
                        for col, val in enumerate(venta_row):
                            self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
                        row_idx += 1

                    else:
                        remaining = q_venta
                        lots = product_lots.setdefault(pid, [])
                        while remaining > 0 and any(safe_int(l["cantidad"]) > 0 for l in lots):
                            if metodo == "UEPS":
                                idx = 0
                                while idx < len(lots) and safe_int(lots[idx]["cantidad"]) <= 0:
                                    idx += 1
                                if idx >= len(lots):
                                    break
                            else:  # PEPS consumes most recent
                                idx = len(lots) - 1
                                while idx >= 0 and safe_int(lots[idx]["cantidad"]) <= 0:
                                    idx -= 1
                                if idx < 0:
                                    break

                            lot = lots[idx]
                            avail = safe_int(lot["cantidad"])
                            take = min(avail, remaining)
                            lot["cantidad"] = avail - take
                            remaining -= take

                            venta_row = [
                                ev["fecha"], "Venta", str(nombre), str(lot["id_inventario"]),
                                "-", "-", "-",
                                str(take), f"{lot['precio']:.2f}", f"{(take * lot['precio']):.2f}",
                                str(lot["cantidad"]), f"{lot['precio']:.2f}", f"{(lot['cantidad'] * lot['precio']):.2f}"
                            ]
                            self.kardex_table.insertRow(row_idx)
                            for col, val in enumerate(venta_row):
                                self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
                            row_idx += 1

                        if remaining > 0:
                            venta_row = [
                                ev["fecha"], "Venta", str(nombre), "-",
                                "-", "-", "-",
                                str(q_venta), f"{ev['precio']:.2f}", f"{(q_venta * ev['precio']):.2f}",
                                str(-remaining), f"{0:.2f}", f"{0:.2f}"
                            ]
                            self.kardex_table.insertRow(row_idx)
                            for col, val in enumerate(venta_row):
                                self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
                            row_idx += 1

            # --- Para PEPS/UEPS: mostrar inventario final POR LOTE y luego UN TOTAL por producto (unidad y suma total) ---
            if metodo in ("PEPS", "UEPS"):
                for pid in sorted(product_lots.keys()):
                    lots = product_lots[pid]
                    cursor.execute("SELECT nombre FROM productos WHERE id_producto = ? LIMIT 1", (pid,))
                    pr = cursor.fetchone()
                    nombre = pr[0] if pr else str(pid)

                    # Mostrar cada lote remanente
                    for lot in lots:
                        if safe_int(lot["cantidad"]) <= 0:
                            continue
                        final_row = [
                            lot["fecha"], "Inventario Final", str(nombre), str(lot["id_inventario"]),
                            "-", "-", "-",
                            "-", "-", "-",
                            str(lot["cantidad"]), f"{lot['precio']:.2f}", f"{(lot['cantidad'] * lot['precio']):.2f}"
                        ]
                        self.kardex_table.insertRow(row_idx)
                        for col, val in enumerate(final_row):
                            self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
                        row_idx += 1

                    # TOTAL por producto sobre todos los lotes (unidades y suma de totales). precio unitario queda vacío.
                    sum_qty = sum(safe_int(l["cantidad"]) for l in lots)
                    if sum_qty > 0:
                        sum_total = sum(safe_int(l["cantidad"]) * safe_float(l["precio"]) for l in lots)
                        total_row = [
                            "-", "TOTAL", str(nombre), "-",
                            "-", "-", "-",
                            "-", "-", "-",
                            str(sum_qty), "", f"{sum_total:.2f}"
                        ]
                        self.kardex_table.insertRow(row_idx)
                        for col, val in enumerate(total_row):
                            self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
                        row_idx += 1

            else:
                # PMP: mostrar TOTAL por producto (cantidad y valoración con precio promedio), ID columna '-'
                for pid in sorted(product_prom.keys()):
                    prog = product_prom[pid]
                    cant = safe_int(prog["cantidad"])
                    if cant <= 0:
                        continue
                    cursor.execute("SELECT nombre FROM productos WHERE id_producto = ? LIMIT 1", (pid,))
                    pr = cursor.fetchone()
                    nombre = pr[0] if pr else str(pid)
                    total_val = cant * safe_float(prog["precio_prom"])
                    total_row = [
                        "-", "TOTAL", str(nombre), "-",
                        "-", "-", "-",
                        "-", "-", "-",
                        str(cant), f"{prog['precio_prom']:.2f}", f"{total_val:.2f}"
                    ]
                    self.kardex_table.insertRow(row_idx)
                    for col, val in enumerate(total_row):
                        self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
                    row_idx += 1

            # Ajustar filas y UI
            self.kardex_table.setRowCount(max(row_idx, 2))
            self.kardex_table.resizeColumnsToContents()
            # volver a limitar ancho ID por si resizeColumns cambió
            try:
                self.kardex_table.setColumnWidth(3, 70)
            except Exception:
                pass
            self.kardex_table.resizeRowsToContents()
            self.kardex_table.verticalHeader().setVisible(False)

    # ================================
    # Exportar a Excel
    # ================================
    def exportar_excel(self):
        if self.kardex_table.rowCount() <= 2:
            QMessageBox.warning(self, "Error", "No hay datos para exportar.")
            return

        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar Kardex en Excel", "", "Archivos Excel (*.xlsx)")
        if not ruta:
            return
        if not ruta.lower().endswith(".xlsx"):
            ruta += ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Kardex"

        # Guardar encabezados (fila 1: columnas) — reemplazar "Precio Unitario" por "Unitario"
        for col in range(self.kardex_table.columnCount()):
            header = self.kardex_table.item(1, col)
            header_text = header.text() if header else ""
            # reemplazo insensible a mayúsculas
            if "precio unitario" in header_text.lower():
                header_text = "Unitario"
            ws.cell(row=1, column=col + 1, value=header_text)

        # Guardar datos (desde fila 2 en adelante)
        out_row = 2
        for row in range(2, self.kardex_table.rowCount()):
            fila_vacia = True
            # check if any cell non-empty
            for col in range(self.kardex_table.columnCount()):
                item = self.kardex_table.item(row, col)
                if item and item.text().strip():
                    fila_vacia = False
                    break
            if fila_vacia:
                continue
            for col in range(self.kardex_table.columnCount()):
                item = self.kardex_table.item(row, col)
                texto = item.text() if item else ""
                ws.cell(row=out_row, column=col + 1, value=texto)
            out_row += 1

        wb.save(ruta)
        QMessageBox.information(self, "Éxito", f"Kardex exportado a:\n{ruta}")

    # ================================
    # Exportar a PDF
    # ================================
    def exportar_pdf(self):
        if self.kardex_table.rowCount() <= 2:
            QMessageBox.warning(self, "Error", "No hay datos para exportar.")
            return

        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar Kardex en PDF", "", "Archivos PDF (*.pdf)")
        if not ruta:
            return
        if not ruta.lower().endswith(".pdf"):
            ruta += ".pdf"

        data = []

        # Encabezados (fila 1 de la tabla UI) — mapear "Precio Unitario" -> "Unitario"
        headers = []
        for col in range(self.kardex_table.columnCount()):
            header = self.kardex_table.item(1, col)
            header_text = header.text() if header else ""
            if "precio unitario" in header_text.lower():
                header_text = "Unitario"
            headers.append(header_text)
            
        data.append(headers)

        # Datos (desde la fila 2 en adelante)
        for row in range(2, self.kardex_table.rowCount()):
            fila = []
            fila_vacia = True
            for col in range(self.kardex_table.columnCount()):
                item = self.kardex_table.item(row, col)
                texto = item.text() if item else ""
                if texto.strip():
                    fila_vacia = False
                fila.append(texto)
            # solo añadir filas que tengan algo
            if not fila_vacia:
                data.append(fila)

        # Usar landscape
        page_size = landscape(letter)
        left_margin = right_margin = top_margin = bottom_margin = 18  # márgenes en puntos

        doc = SimpleDocTemplate(ruta, pagesize=page_size,
                                leftMargin=left_margin, rightMargin=right_margin,
                                topMargin=top_margin, bottomMargin=bottom_margin)

        elementos = []
        styles = getSampleStyleSheet()
        elementos.append(Paragraph("Reporte de Kardex", styles["Heading1"]))
        elementos.append(Spacer(1, 12))

        # calcular ancho util (ancho de página menos márgenes)
        page_width, page_height = page_size
        usable_width = page_width - left_margin - right_margin

        # distribuir columnas uniformemente
        col_count = max(1, self.kardex_table.columnCount())
        col_widths = [usable_width / col_count] * col_count

        tabla = Table(data, colWidths=col_widths, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        elementos.append(tabla)
        doc.build(elementos)

        QMessageBox.information(self, "Éxito", f"Kardex exportado a:\n{ruta}")
