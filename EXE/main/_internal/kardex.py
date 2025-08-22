from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem,
    QPushButton, QMessageBox, QHBoxLayout, QDateEdit, QFileDialog
)
from PyQt6.QtGui import QIcon
from PyQt6.QtCore import QDate, Qt
import sqlite3
import database as db
import os
import sys

# Exportaciones
from openpyxl import Workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet


def safe_int(v):
    try:
        return int(v)
    except:
        return 0

def safe_float(v):
    try:
        return float(v)
    except:
        return 0.0

def resource_path(relative_path):
    """Obtiene la ruta absoluta del recurso, compatible con PyInstaller."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class KardexWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Kardex de Ventas")
        self.setMinimumSize(1100, 650)
        self.setWindowIcon(QIcon(resource_path("mainlogo.ico")))

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout()

        # Selección de rango de fechas
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("Fecha inicio:"))
        self.fecha_inicio = QDateEdit()
        self.fecha_inicio.setCalendarPopup(True)
        self.fecha_inicio.setDate(QDate.currentDate().addDays(-30))
        date_layout.addWidget(self.fecha_inicio)

        date_layout.addWidget(QLabel("Fecha fin:"))
        self.fecha_fin = QDateEdit()
        self.fecha_fin.setCalendarPopup(True)
        self.fecha_fin.setDate(QDate.currentDate())
        date_layout.addWidget(self.fecha_fin)

        layout.addLayout(date_layout)

        # Botón mostrar Kardex
        btn_mostrar = QPushButton("Mostrar Kardex")
        btn_mostrar.clicked.connect(self.mostrar_kardex)
        layout.addWidget(btn_mostrar)

        # Botones exportación
        export_layout = QHBoxLayout()
        btn_excel = QPushButton("Exportar a Excel")
        btn_excel.clicked.connect(self.exportar_excel)
        export_layout.addWidget(btn_excel)

        btn_pdf = QPushButton("Exportar a PDF")
        btn_pdf.clicked.connect(self.exportar_pdf)
        export_layout.addWidget(btn_pdf)

        layout.addLayout(export_layout)

        # Tabla Kardex
        self.kardex_table = QTableWidget()
        layout.addWidget(self.kardex_table)

        central.setLayout(layout)


        
    def mostrar_kardex(self):
        conn = sqlite3.connect(db.DB_NAME)
        cursor = conn.cursor()

        fecha_inicio = self.fecha_inicio.date().toString("yyyy-MM-dd")
        fecha_fin = self.fecha_fin.date().toString("yyyy-MM-dd")

        # ================================
        # Movimientos con liberaciones
        # ================================
        cursor.execute("""
            SELECT
                l.fecha AS fecha_liberacion,
                l.id_liberacion,
                v.id_venta,
                p.nombre AS producto,
                c.id_compra AS id_compra,
                dc.cantidad AS compra_cantidad,
                dc.precio_unitario AS compra_precio,
                (dc.cantidad * dc.precio_unitario) AS compra_total,
                li.cantidad AS salida_cantidad,
                i.precio_unitario AS salida_precio,
                (li.cantidad * i.precio_unitario) AS salida_total
            FROM liberaciones l
            JOIN liberacion_inventarios li ON li.id_liberacion = l.id_liberacion
            JOIN inventarios i ON i.id = li.id_inventario
            JOIN compras c ON c.id_compra = i.id_compra
            JOIN detalle_compras dc ON dc.id_compra = c.id_compra AND dc.id_producto = i.id_producto
            JOIN productos p ON p.id_producto = i.id_producto
            LEFT JOIN detalle_ventas dv ON dv.id_venta = l.id_venta AND dv.id_producto = i.id_producto
            LEFT JOIN ventas v ON v.id_venta = l.id_venta
            WHERE l.fecha BETWEEN ? AND ?
            ORDER BY l.fecha ASC, l.id_liberacion ASC, i.id ASC
        """, (fecha_inicio, fecha_fin))
        filas = cursor.fetchall()

        # ================================
        # Todas las compras (para visualización y total)
        # ================================
        cursor.execute("""
            SELECT
                c.fecha AS fecha_compra,
                p.nombre AS producto,
                c.id_compra AS id_compra,
                dc.cantidad AS compra_cantidad,
                dc.precio_unitario AS compra_precio,
                (dc.cantidad * dc.precio_unitario) AS compra_total
            FROM compras c
            JOIN detalle_compras dc ON dc.id_compra = c.id_compra
            JOIN productos p ON p.id_producto = dc.id_producto
            WHERE c.fecha BETWEEN ? AND ?
            ORDER BY c.fecha ASC, c.id_compra ASC
        """, (fecha_inicio, fecha_fin))
        todas_compras = cursor.fetchall()

        if not filas and not todas_compras:
            QMessageBox.information(self, "Info", "No hay movimientos en el rango de fechas.")
            self.kardex_table.clear()
            self.kardex_table.setRowCount(0)
            self.kardex_table.setColumnCount(0)
            conn.close()
            return

        columnas = [
            "Fecha", "Tipo", "Producto",
            "Cantidad", "Precio Unitario", "Valor Total",
            "Cantidad", "Precio Unitario", "Valor Total",
            "Cantidad", "Precio Unitario", "Valor Total"
        ]
        grupos = [
            "", "", "",
            "Entradas", "Entradas", "Entradas",
            "Salidas", "Salidas", "Salidas",
            "Inventario Final", "Inventario Final", "Inventario Final"
        ]

        total_rows = (len(filas) * 2) + len(todas_compras)
        self.kardex_table.setColumnCount(len(columnas))
        self.kardex_table.setRowCount(total_rows + 10)

        # Encabezados
        for col, texto in enumerate(grupos):
            item = QTableWidgetItem(texto)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.kardex_table.setItem(0, col, item)

        for col, texto in enumerate(columnas):
            item = QTableWidgetItem(texto)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.kardex_table.setItem(1, col, item)

        self.kardex_table.setSpan(0, 0, 2, 1)
        self.kardex_table.setSpan(0, 1, 2, 1)
        self.kardex_table.setSpan(0, 2, 2, 1)

        def merge_group(start_col, span, text):
            self.kardex_table.setSpan(0, start_col, 1, span)
            itm = self.kardex_table.item(0, start_col)
            if itm:
                itm.setText(text)
            else:
                self.kardex_table.setItem(0, start_col, QTableWidgetItem(text))

        merge_group(3, 3, "Entradas")
        merge_group(6, 3, "Salidas")
        merge_group(9, 3, "Inventario Final")

        def safe_int(v):
            try:
                return int(v)
            except:
                return 0

        def safe_float(v):
            try:
                return float(v)
            except:
                return 0.0

        inventario_actual = {}   # por id_compra -> (cantidad_restante, precio)
        inventario_total = {}    # por producto -> cantidad total
        row_idx = 2

        # ================================
        # Mostrar todas las compras (estéticas)
        # ================================
        for (fecha_compra, producto, id_compra, compra_cantidad, compra_precio, compra_total) in todas_compras:
            compra_cant = safe_int(compra_cantidad)
            compra_prec = safe_float(compra_precio)
            compra_tot = safe_float(compra_total)
            inventario_total[producto] = inventario_total.get(producto, 0) + compra_cant

            compra_row = [
                str(fecha_compra), "Compra", str(producto),
                str(compra_cant), f"{compra_prec:.2f}", f"{compra_tot:.2f}",
                "-", "-", "-",
                str(compra_cant), f"{compra_prec:.2f}", f"{compra_tot:.2f}"
            ]
            for col, val in enumerate(compra_row):
                self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
            row_idx += 1

        # ================================
        # Procesar ventas, copiando la compra antes de cada venta con stock real
        # ================================
        for (fecha_lib, id_liberacion, id_venta, producto, id_compra,
            compra_cantidad, compra_precio, compra_total,
            salida_cantidad, salida_precio, salida_total) in filas:

            compra_cant = safe_int(compra_cantidad)
            compra_prec = safe_float(compra_precio)
            compra_tot = safe_float(compra_total)

            # Obtener stock real antes de la venta
            stock_cant, stock_prec = inventario_actual.get(id_compra, (compra_cant, compra_prec))

            # Copiar compra antes de la venta (estética) usando stock real
            compra_row = [
                str(fecha_lib), "Compra", str(producto),
                str(stock_cant), f"{stock_prec:.2f}", f"{stock_cant * stock_prec:.2f}",
                "-", "-", "-",
                str(stock_cant), f"{stock_prec:.2f}", f"{stock_cant * stock_prec:.2f}"
            ]
            for col, val in enumerate(compra_row):
                self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
            row_idx += 1

            # Registrar stock real
            if id_compra not in inventario_actual:
                inventario_actual[id_compra] = (compra_cant, compra_prec)

            salida_cant = safe_int(salida_cantidad)
            salida_prec = safe_float(salida_precio)
            salida_tot = safe_float(salida_total)

            stock_cant, stock_prec = inventario_actual.get(id_compra, (0, salida_prec))
            stock_cant -= salida_cant
            if stock_cant < 0:
                stock_cant = 0
            inventario_actual[id_compra] = (stock_cant, stock_prec)

            inventario_total[producto] = max(inventario_total.get(producto, 0) - salida_cant, 0)

            final_row = ["-", "-", "-"] if stock_cant <= 0 else [
                str(stock_cant), f"{stock_prec:.2f}", f"{stock_cant * stock_prec:.2f}"
            ]

            venta_row = [
                str(fecha_lib), "Venta", str(producto),
                "-", "-", "-",
                str(salida_cant) if salida_cant else "-",
                f"{salida_prec:.2f}" if salida_cant else "-",
                f"{salida_tot:.2f}" if salida_cant else "-",
                final_row[0], final_row[1], final_row[2]
            ]
            for col, val in enumerate(venta_row):
                self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
            row_idx += 1

        # ================================
        # TOTAL por producto (solo cada compra una vez)
        # ================================
            for producto in set([p for _, p, *_ in todas_compras]):
                total_cantidad = 0
                total_valor = 0.0

                # Recorremos todas las compras del producto
                for (fecha_compra, prod, id_compra, cant, prec, _) in todas_compras:
                    if prod != producto:
                        continue

                    # Ver cuánto queda realmente de esta compra
                    if id_compra in inventario_actual:
                        cant_rest, prec_rest = inventario_actual[id_compra]
                        if cant_rest > 0:
                            total_cantidad += cant_rest
                            total_valor += cant_rest * prec_rest
                    else:
                        # Compra nunca usada → queda todo
                        cant_ini = safe_int(cant)
                        prec_ini = safe_float(prec)
                        total_cantidad += cant_ini
                        total_valor += cant_ini * prec_ini

                if total_cantidad > 0:
                    total_row = [
                        "-", "TOTAL", producto,
                        "-", "-", "-",
                        "-", "-", "-",
                        str(total_cantidad), "-", f"{total_valor:.2f}"
                    ]
                    for col, val in enumerate(total_row):
                        self.kardex_table.setItem(row_idx, col, QTableWidgetItem(str(val)))
                    row_idx += 1
                    
        conn.close()
        self.kardex_table.resizeColumnsToContents()
        self.kardex_table.resizeRowsToContents()
        self.kardex_table.verticalHeader().setVisible(False)

    # ================================
    # Exportar a Excel
    # ================================
    def exportar_excel(self):
        if self.kardex_table.rowCount() == 0:
            QMessageBox.warning(self, "Error", "No hay datos para exportar.")
            return

        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar Kardex en Excel", "", "Archivos Excel (*.xlsx)")
        if not ruta:
            return  # Usuario canceló

        wb = Workbook()
        ws = wb.active
        ws.title = "Kardex"

        # Guardar encabezados
        for col in range(self.kardex_table.columnCount()):
            header = self.kardex_table.item(1, col)
            if header:
                ws.cell(row=1, column=col+1, value=header.text())

        # Guardar datos
        for row in range(2, self.kardex_table.rowCount()):
            for col in range(self.kardex_table.columnCount()):
                item = self.kardex_table.item(row, col)
                if item:
                    ws.cell(row=row, column=col+1, value=item.text())

        wb.save(ruta)
        QMessageBox.information(self, "Éxito", f"Kardex exportado a:\n{ruta}")


    # ================================
    # Exportar a PDF
    # ================================
    def exportar_pdf(self):
        if self.kardex_table.rowCount() == 0:
            QMessageBox.warning(self, "Error", "No hay datos para exportar.")
            return

        ruta, _ = QFileDialog.getSaveFileName(self, "Guardar Kardex en PDF", "", "Archivos PDF (*.pdf)")
        if not ruta:
            return  # Usuario canceló

        data = []

        # Encabezados
        headers = []
        for col in range(self.kardex_table.columnCount()):
            header = self.kardex_table.item(1, col)
            headers.append(header.text() if header else "")
        data.append(headers)

        # Datos
        for row in range(2, self.kardex_table.rowCount()):
            fila = []
            for col in range(self.kardex_table.columnCount()):
                item = self.kardex_table.item(row, col)
                fila.append(item.text() if item else "")
            data.append(fila)

        doc = SimpleDocTemplate(ruta, pagesize=letter)
        elementos = []

        styles = getSampleStyleSheet()
        elementos.append(Paragraph("Reporte de Kardex", styles["Heading1"]))
        elementos.append(Spacer(1, 12))

        tabla = Table(data)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))

        elementos.append(tabla)
        doc.build(elementos)

        QMessageBox.information(self, "Éxito", f"Kardex exportado a:\n{ruta}")